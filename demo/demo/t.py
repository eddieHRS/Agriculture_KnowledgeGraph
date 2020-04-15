import openpyxl
import os

os.chdir("C:/Users/eddie_00z60yf/Desktop/test")

from openpyxl import load_workbook

filename = 'a.xlsx'
filename2 = 'b.xlsx'
wb = load_workbook(filename)
wb2 = load_workbook(filename2)
from_sheet_names = wb.sheetnames
to_sheet_names = wb2.sheetnames
print(from_sheet_names, to_sheet_names)

def replace_xls(from_sheet_name, to_sheet_name):
    ws = wb[from_sheet_name]
    ws2 = wb2[to_sheet_name]
    # 两个for循环遍历整个excel的单元格内容
    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            ws2.cell(row=i + 1, column=j + 1, value=cell.value)

    wb2.save(filename2)

# 遇到复制几十个sheet时候，很有必要写个循环
for i in range(0,len(from_sheet_names)):
    replace_xls(from_sheet_names[i], to_sheet_names[i])